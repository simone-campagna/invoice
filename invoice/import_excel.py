"""
Import from excel
"""

import argparse
import collections
import datetime
import os
import re

from openpyxl import load_workbook

from . import conf

Field = collections.namedtuple(
    'Field', 'header field type')


OUT_DATE_FMT = '%d/%m/%Y'


def mk_vat_number(x):
    if x:
        return x.strip()
    else:
        return None


def mk_date(x):
    return datetime.date.fromordinal(datetime.date(1899,12, 30).toordinal() + x).strftime(OUT_DATE_FMT)


def mk_p_vat(x):
    if x.strip() in {'A1', 'A10_18'}:
        return  mk_float(0)
    else:
        return  mk_float(x.strip())


def mk_float(x):
    return float(x)


_EXCEPTIONS = [
    (re.compile(r"\s+non\s+.*addebito\s+.*\s+marca\s+da\s+bollo"), 'no-bollo'),
]


def mk_exceptions(x):
    exceptions = []
    if x:
        for r, key in _EXCEPTIONS:
            if r.search(x):
                exceptions.append(key)
    return ','.join(exceptions)


def read_workbook(filename, fields):
    def strip(txt):
        return txt.replace(' ', '')

    wb = load_workbook(filename)
    if len(wb.sheetnames) != 1:
        raise ValueError("file {}: sheetnames: {!r}".format(filename, wb.sheetnames))
    sheetname = wb.sheetnames[0]
    ws = wb[sheetname]
    iws = iter(ws.rows)
    header = [strip(cell.value) for cell in next(iws)]
    fields_dict = {strip(field.header): field for field in fields}
    cols = {}
    for index, value in enumerate(header):
        field = fields_dict.pop(value, None)
        if field is not None:
            cols[index] = field
    for field in fields_dict.values():
        raise ValueError("{}: field {} not found".format(filename, field))
    # for index, field in sorted(cols.items(), key=lambda x: x[0]):
    #     hdr = str(header[index])
    #     if hdr != field.header:
    #         raise ValueError("file {}: col {}={!r} is not {!r}".format(filename, index, hdr, field.header))
    for row in iws:
        dct = {}
        for index, field in cols.items():
            # print(index, field, row)
            dct[field.field] = field.type(row[index].value)
        yield dct


def read_clients(filename):
    fields = [
        Field('Cliente', 'name', str),
        Field('Indirizzo', 'address', str),
        Field('Comune', 'city', str),
        Field('CodiceFiscale', 'tax_code', str),
    ]
    clients = {}
    for dct in read_workbook(filename, fields):
        if dct['name'] in clients:
            raise ValueError("client {!r} already defined".format(dct['name']))
        clients[dct['tax_code']] = dct
    return clients


def create_document(filename, year, number, rows, clients):
    if len(rows) != 1:
        raise ValueError("unsupported number of entries #{} in invoice {}/{:03d}".format(len(rows), year, number))
    filename = filename.format(year=year, number=number)
    dirname = os.path.dirname(os.path.abspath(filename))
    if not os.path.isdir(dirname):
        os.makedirs(dirname)
    taxable_income = 0
    fee = 0
    for row in rows:
        taxable_income += sum(row.get(key, 0) for key in ['fee', 'cpa', 'refunds'])
        fee += row['fee']
    # REM if taxable_income > 77.47 and row['vat'] == 0:
    # REM     taxes = 2.0
    # REM else:
    # REM     taxes = 0.0
    # print(taxable_income, row['vat'], row['deduction'], taxes)

    template = """\
Fattura n° {year}/{number:03d}
						Casalecchio di Reno, {date}
                                                    Spett. {name}
{address}
12345 {city}
							Cod.Fisc. {tax_code}
                     
FAC SIMILE

N° 1 {service}	{fee} euro

Rimborso spese di viaggio	{refunds} euro
Contributo previdenziale {p_cpa}%	{cpa} euro
IVA {p_vat}%	{vat} euro
Ritenuta d'acconto {p_deduction}%	{deduction} euro
Bollo (abc)	{taxes} euro
Totale fattura	{income} euro

################################################################################

# year_and_number|{year}|{number}

# name|{name}

# tax_code|{tax_code}

# city_and_date|{city}|{date}

# income_and_currency|{income}|euro

# service_and_fee|{service}|{fee}

# p_vat_and_vat|{p_vat}|{vat}

# p_deduction_and_deduction|{p_deduction}|{deduction}

# p_cpa_and_cpa|{p_cpa}|{cpa}

# refunds|{refunds}

# taxes|{taxes}

# exceptions|{exceptions}
"""
    ref_row = rows[-1]
    vat_number = None
    for key in 'e_vat_number', 'p_vat_number':
        if ref_row[key] is not None:
            vat_number = ref_row[key]
            break
    else:
        raise ValueError("{}: c.f./p.iva mancante".format(ref_row))
    if vat_number not in clients:
        raise ValueError("cliente {!r} non in anagrafica".format(vat_number))
    client = clients[vat_number]
    data = ref_row.copy()
    data.update({
        'refunds': 0.0,
        'fee': fee,
        # REM 'taxes': taxes,
        'address': client['address'],
        'tax_code': client['tax_code'],
        'city': client['city'],
    })
    def conv(k, v):
        if k in {'fee', 'cpa', 'vat', 'refunds', 'deduction', 'taxes', 'income'}:
            return str(v).replace('.', ',')
        else:
            return v
        
    data['income'] += data['deduction']
    test = template.format(**{k: conv(k, v) for k, v in data.items()})

    with open(filename, "w") as o_file:
        o_file.write(test)

def create_documents(clients, invoices, filename):
    for (year, number), rows in invoices.items():
        create_document(filename, year, number, rows, clients)

def read_invoices(filename):
    fields = [
        Field('Anno', 'year', int),
        Field('Numero', 'number', int),
        Field('Data', 'date', mk_date),
        Field('Cliente/Fornitore', 'name', str),
        Field('C.F.', 'p_vat_number', mk_vat_number),
        Field('P.I.', 'e_vat_number', mk_vat_number),
        Field('Numero riga', 'num_row', int),
        Field('Descrizione', 'service', str),
        Field('PrezzoTot', 'fee', mk_float),
        Field('Aliquota', 'p_vat', mk_p_vat),
        Field('Totale (val)', 'income', mk_float),
        Field('Imposta (val)', 'vat', mk_float),
        Field('Cassa Previdenza (%)', 'p_cpa', mk_float),
        Field('Cassa previdenza (val)', 'cpa', mk_float),
        Field('Ritenuta (%)', 'p_deduction', mk_float),
        Field('Ritenuta (val)', 'deduction', mk_float),
        Field('Note piede', 'exceptions', mk_exceptions),
        Field('Bollo (val)', 'taxes', mk_float),
    ]
    invoices = collections.defaultdict(list)
    for dct in read_workbook(filename, fields):
        invoices[(dct['year'], dct['number'])].append(dct)
    return invoices


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--clients", "-c",
        required=True,
        type=str,
        help="anagrafica clienti")
    parser.add_argument(
        "--invoices", "-i",
        required=True,
        type=str,
        help="fatture")
    parser.add_argument(
        "--output", "-o",
        default="test-docs/{year}_{number}.doc",
        type=str,
        help="documento di output")

    namespace = parser.parse_args()

    # lettura anagrafica:
    clients = read_clients(namespace.clients)

    # lettura fatture:
    invoices = read_invoices(namespace.invoices)

    # creazione documenti:
    for (year, number), rows in invoices.items():
        create_document(namespace.output, year, number, rows, clients)


if __name__ == "__main__":
    main()
    
